"""Equivalence tests for the streaming pipeline refactor.

Golden CSV/JSON/XLSX in tests/golden/ were produced by the pre-refactor code
on the fixture tests/fixtures/sample_small.log (~2000 zapp lines extracted
from a real ZPA syslog file). Any change to the parsing/session pipeline
must keep these outputs identical (CSV byte-for-byte, JSON dict-equal,
XLSX values identical).
"""

import inspect
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime
from types import GeneratorType

import pytest
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

from session_parser import (
    REPORT_COLUMNS,
    _SessionAcc,
    build_sessions,
    iter_log_records,
    parse_log_file,
)

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
GOLDEN_DIR = os.path.join(os.path.dirname(__file__), "golden")
SAMPLE_LOG = os.path.join(FIXTURES_DIR, "sample_small.log")
GOLDEN_CSV = os.path.join(GOLDEN_DIR, "sample_small.csv")
GOLDEN_JSON = os.path.join(GOLDEN_DIR, "sample_small.json")
GOLDEN_XLSX = os.path.join(GOLDEN_DIR, "sample_small.xlsx")

REPORT_DATE = "2026-04-19"
TIMEZONE = "Europe/Rome"


def _run_generator(tmp_path):
    """Invoke report_generator.py as a subprocess against the fixture.

    Subprocess (vs in-process) isolates module state (logging singletons,
    sys.argv) and matches how the systemd timer runs it in production.
    """
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    generator = os.path.join(project_root, "src", "report_generator.py")
    out_dir = str(tmp_path)
    env = os.environ.copy()
    env["PYTHONPATH"] = os.path.join(project_root, "src")
    env["ZPA_SIEM_CONFIG"] = os.path.join(project_root, "config.ini")
    result = subprocess.run(
        [
            sys.executable, generator,
            "--log-file", SAMPLE_LOG,
            "--output-dir", out_dir,
            "--date", REPORT_DATE,
            "--timezone", TIMEZONE,
            "--no-upload",
        ],
        capture_output=True, text=True, env=env, timeout=60,
    )
    assert result.returncode == 0, (
        f"generator failed (exit {result.returncode}):\n{result.stdout}\n{result.stderr}"
    )
    return {
        "csv": os.path.join(out_dir, f"zpa-report-{REPORT_DATE}.csv"),
        "json": os.path.join(out_dir, f"zpa-report-{REPORT_DATE}.json"),
        "xlsx": os.path.join(out_dir, f"zpa-report-{REPORT_DATE}.xlsx"),
    }


def test_csv_byte_identical_to_golden(tmp_path):
    """CSV output must be byte-for-byte identical to the golden file."""
    outputs = _run_generator(tmp_path)
    with open(outputs["csv"], "rb") as f:
        new_bytes = f.read()
    with open(GOLDEN_CSV, "rb") as f:
        golden_bytes = f.read()
    assert new_bytes == golden_bytes


def test_json_dict_equal_to_golden(tmp_path):
    """JSON output must be dict-equal to the golden file, ignoring generated_at.

    generated_at is a wall-clock timestamp that differs every run.
    """
    outputs = _run_generator(tmp_path)
    with open(outputs["json"]) as f:
        new = json.load(f)
    with open(GOLDEN_JSON) as f:
        golden = json.load(f)
    new.pop("generated_at", None)
    golden.pop("generated_at", None)
    assert new == golden


def test_xlsx_values_identical_to_golden(tmp_path):
    """XLSX cell values (header + data rows) must match the golden file."""
    outputs = _run_generator(tmp_path)
    new_rows = _xlsx_rows(outputs["xlsx"])
    golden_rows = _xlsx_rows(GOLDEN_XLSX)
    assert new_rows == golden_rows


def _xlsx_rows(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def test_iter_log_records_returns_generator():
    """iter_log_records must yield, not materialise a list."""
    it = iter_log_records(SAMPLE_LOG)
    assert isinstance(it, GeneratorType) or inspect.isgenerator(it)


def test_iter_log_records_prefilters_zapp_only():
    """iter_log_records yields only zpn_client_type_zapp records.

    The fixture was extracted with `grep '"ClientType": "zpn_client_type_zapp"'`,
    so every line is already a zapp record — count must match.
    """
    count = sum(1 for _ in iter_log_records(SAMPLE_LOG))
    with open(SAMPLE_LOG) as f:
        line_count = sum(1 for line in f if line.strip())
    assert count == line_count


def test_parse_log_file_legacy_still_returns_list():
    """The legacy parse_log_file entry point must still return a list (no breaking change)."""
    recs = parse_log_file(SAMPLE_LOG)
    assert isinstance(recs, list)
    assert all(isinstance(r, dict) for r in recs)
    assert len(recs) > 0


def test_streaming_vs_list_input_match():
    """build_sessions must produce identical output whether fed a list or a generator."""
    tz = ZoneInfo(TIMEZONE)
    list_recs = list(iter_log_records(SAMPLE_LOG))

    sessions_from_list = build_sessions(list_recs, tz, max_client_version=10)
    sessions_from_gen = build_sessions(iter_log_records(SAMPLE_LOG), tz, max_client_version=10)

    assert len(sessions_from_list) == len(sessions_from_gen)
    # Sorted by (Username, Date, Session Start) inside build_sessions, so order
    # is deterministic across both inputs.
    for a, b in zip(sessions_from_list, sessions_from_gen):
        assert a == b


def test_session_acc_uses_slots():
    """_SessionAcc must declare __slots__ to keep the per-SID footprint small."""
    assert hasattr(_SessionAcc, "__slots__")
    assert not hasattr(_SessionAcc(_dummy_first_rec(), datetime.now(ZoneInfo("UTC"))), "__dict__"), (
        "_SessionAcc instances must not carry a __dict__ (slots violation)"
    )


def _dummy_first_rec():
    return {
        "Username": "x",
        "TimestampAuthentication": "2026-01-01T00:00:00.000Z",
    }


def test_report_columns_unchanged():
    """REPORT_COLUMNS is part of the output contract — guard against accidental drift."""
    assert REPORT_COLUMNS == [
        "Username", "Date", "Session Start", "Session End", "Duration",
        "Main Public IP", "Other Public IPs", "Main Private IP", "Other Private IPs",
        "City", "Country", "Device", "Platform", "Client Version",
        "Trusted Network", "Bytes Rx", "Bytes Tx",
    ]

#!/usr/bin/env python3
"""ZPA Status Mini-SIEM — Daily Report Generator.

Reads ZPA syslog files, extracts user sessions, and generates Excel + JSON reports.

Usage:
    python3 report_generator.py                        # process yesterday's log
    python3 report_generator.py --date 2026-04-09      # process a specific date
    python3 report_generator.py --log-file path.log    # process a specific file
    python3 report_generator.py --output-dir ./reports  # custom output directory
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from session_parser import REPORT_COLUMNS, build_sessions, iter_log_records, parse_timestamp
from config import Config
from app_logger import setup_logging, get_logger

log = get_logger(__name__)


# --- Excel generation ---


def generate_excel(sessions: list, output_path: str) -> None:
    """Generate the Excel report from consolidated sessions.

    Uses openpyxl's write-only mode: rows are streamed straight to the XLSX
    archive without keeping the whole sheet in memory, and column widths are
    accumulated during the single append pass (replacing the legacy second
    iter_rows traversal). Output is functionally identical to the legacy
    generator: same headers, same styles, same freeze_panes and auto_filter.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="ZPA Sessions")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    active_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    # Pre-build per-column data alignments (one object reused per column
    # across all data rows — cheaper than constructing one per cell).
    data_alignments = [
        Alignment(horizontal="center" if col_name != "Username" else "left")
        for col_name in REPORT_COLUMNS
    ]

    max_widths = [len(c) for c in REPORT_COLUMNS]

    header_cells = []
    for col_name in REPORT_COLUMNS:
        c = WriteOnlyCell(ws, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
        c.border = thin_border
        header_cells.append(c)
    ws.append(header_cells)

    row_count = 0
    for session in sessions:
        is_active = session["Session End"] == "In corso"
        row_cells = []
        for i, col_name in enumerate(REPORT_COLUMNS):
            value = session[col_name]
            c = WriteOnlyCell(ws, value=value)
            c.border = thin_border
            c.alignment = data_alignments[i]
            if is_active:
                c.fill = active_fill
            row_cells.append(c)
            if value:
                ln = len(str(value))
                if ln > max_widths[i]:
                    max_widths[i] = ln
        ws.append(row_cells)
        row_count += 1

    for i, w in enumerate(max_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}{row_count + 1}"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


# --- JSON generation ---


def generate_csv(sessions: list[dict], output_path: str) -> None:
    """Generate a semicolon-delimited CSV report from consolidated sessions."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f, fieldnames=REPORT_COLUMNS, delimiter=";", extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(sessions)


def _session_to_json_row(s: dict) -> dict:
    return {
        "username": s["Username"],
        "date": s["Date"],
        "session_start": s["Session Start"],
        "session_end": s["Session End"],
        "duration": s["Duration"],
        "main_public_ip": s["Main Public IP"],
        "other_public_ips": [ip for ip in s["Other Public IPs"].split(", ") if ip],
        "main_private_ip": s["Main Private IP"],
        "other_private_ips": [ip for ip in s["Other Private IPs"].split(", ") if ip],
        "city": s["City"],
        "country": s["Country"],
        "device": s["Device"],
        "platform": s["Platform"],
        "client_version": s["Client Version"],
        "trusted_network": s["Trusted Network"],
        "bytes_rx": s["Bytes Rx"],
        "bytes_tx": s["Bytes Tx"],
        "session_ids": s.get("_session_ids", []),
    }


def generate_json(sessions: list, output_path: str, timezone_name: str) -> None:
    """Generate the JSON report from consolidated sessions.

    Writes incrementally: one session row is serialised and flushed at a time
    instead of building the full report dict in memory. Output layout matches
    the legacy `json.dump(report, f, indent=2)` exactly (2-space indent).
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    report_date = sessions[0]["Date"] if sessions else ""
    generated_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("{\n")
        f.write(f'  "report_date": {json.dumps(report_date, ensure_ascii=False)},\n')
        f.write(f'  "timezone": {json.dumps(timezone_name, ensure_ascii=False)},\n')
        f.write(f'  "generated_at": {json.dumps(generated_at, ensure_ascii=False)},\n')
        if not sessions:
            f.write('  "sessions": []\n}')
            return
        f.write('  "sessions": [\n')
        last_idx = len(sessions) - 1
        for i, s in enumerate(sessions):
            row = _session_to_json_row(s)
            # Serialise the row at top-level indent=2, then shift every line
            # right by 4 spaces so it sits two levels deep inside the document
            # (matching what json.dump(report, indent=2) would have produced).
            row_str = json.dumps(row, indent=2, ensure_ascii=False)
            indented = "\n".join("    " + line for line in row_str.split("\n"))
            f.write(indented)
            f.write(",\n" if i != last_idx else "\n")
        f.write("  ]\n}")


def generate_summary_sidecar(sessions: list, output_path: str,
                             report_date: str,
                             has_csv: bool, has_xlsx: bool) -> None:
    """Write a small `{date}.summary.json` next to the full report.

    The dashboard reads sidecars to render the report index and to pre-filter
    candidate dates for username search without ever loading the full JSON
    payload (which can be hundreds of MB at 15k-user scale). Schema is kept
    minimal: only what the dashboard actually consumes.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    usernames = sorted({s["Username"] for s in sessions if s.get("Username")})
    summary = {
        "date": report_date,
        "session_count": len(sessions),
        "has_csv": has_csv,
        "has_xlsx": has_xlsx,
        "usernames": usernames,
        "generated_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False)


# --- Retention cleanup ---


def cleanup_old_reports(output_dir: str, retention_days: int) -> None:
    """Delete report files older than retention_days."""
    if retention_days <= 0:
        return
    cutoff = datetime.now() - timedelta(days=retention_days)
    date_pattern = re.compile(r"(\d{4}-\d{2}-\d{2})")

    for filename in os.listdir(output_dir):
        if not (filename.endswith(".xlsx") or filename.endswith(".csv") or filename.endswith(".json")):
            continue
        match = date_pattern.search(filename)
        if not match:
            continue
        try:
            file_date = datetime.strptime(match.group(1), "%Y-%m-%d")
        except ValueError:
            continue
        if file_date < cutoff:
            filepath = os.path.join(output_dir, filename)
            os.remove(filepath)
            log.info("Cleaned up old report: %s", filename)


# --- Main ---


def find_log_sources(log_dir: str, report_date: str = None) -> tuple[list[str], str]:
    """Find all log files that may contain data for the given date.

    logrotate does NOT rotate at midnight — it runs when the system
    cron fires (typically ~3 AM on RHEL).  So a day's data (midnight
    to midnight) is almost always split across two files:

      - The rotated file from that day's rotation (contains the early
        morning hours: 00:00 → ~03:00)
      - The rotated file from the next day's rotation, OR the active
        zpa.log (contains the rest: ~03:00 → 23:59)

    We collect ALL candidate files and let the caller date-filter.

    Args:
        log_dir: directory containing log files
        report_date: target date as YYYY-MM-DD (default: yesterday)

    Returns (log_paths, report_date).
    """
    if report_date:
        target = datetime.strptime(report_date, "%Y-%m-%d")
    else:
        target = datetime.now() - timedelta(days=1)
        report_date = target.strftime("%Y-%m-%d")

    target_fmt = target.strftime("%Y%m%d")
    next_day_fmt = (target + timedelta(days=1)).strftime("%Y%m%d")

    # Rotated file from the day after target (contains bulk of target day)
    # and rotated file from target day (contains early morning of target)
    candidates = [
        f"zpa.log-{next_day_fmt}",
        f"zpa.log-{next_day_fmt}.gz",
        f"zpa.log-{target_fmt}",
        f"zpa.log-{target_fmt}.gz",
    ]

    found = []
    for name in candidates:
        path = os.path.join(log_dir, name)
        if os.path.exists(path):
            found.append(path)

    # Include the active log only if target date is recent (today or yesterday)
    # — for older dates the active log cannot contain relevant data
    today = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if report_date in (today, yesterday):
        active = os.path.join(log_dir, "zpa.log")
        if os.path.exists(active):
            found.append(active)

    return found, report_date


def main():
    config = Config()

    parser = argparse.ArgumentParser(description="ZPA Status Mini-SIEM Report Generator")
    parser.add_argument("--log-file", help="Path to a specific log file to process")
    parser.add_argument("--log-dir", default=config.log_dir, help="Directory containing log files")
    parser.add_argument("--date", default=None, help="Report date YYYY-MM-DD (default: yesterday)")
    parser.add_argument("--output-dir", default=config.output_dir, help="Directory for output reports")
    parser.add_argument("--output-file", help="Specific output file path (overrides --output-dir)")
    parser.add_argument("--timezone", default=config.timezone_name, help="Timezone for report timestamps")
    parser.add_argument("--config", default=None, help="Path to config.ini file")
    parser.add_argument("--no-upload", action="store_true",
                        help="Skip share upload even if enabled in config (used by regen)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and build sessions but do NOT write report files or upload. "
                             "For production smoke tests.")
    args = parser.parse_args()

    # Reload config if custom path specified
    if args.config:
        config = Config(args.config)

    setup_logging(config.app_log_dir, config.app_log_file)

    started = time.monotonic()
    tz = ZoneInfo(args.timezone)
    log.info("Report run started (timezone=%s)", args.timezone)

    # Validate --date format if provided
    if args.date:
        try:
            datetime.strptime(args.date, "%Y-%m-%d")
        except ValueError:
            log.error("Invalid date format: %s (expected YYYY-MM-DD)", args.date)
            sys.exit(1)

    # Determine input file(s) and report date
    if args.log_file:
        log_paths = [args.log_file]
        # For manual runs with --log-file, derive date from filename or --date or today
        if args.date:
            report_date = args.date
        else:
            date_match = re.search(r"(\d{4})(\d{2})(\d{2})", os.path.basename(args.log_file))
            if date_match:
                y, m, d = date_match.groups()
                report_date = f"{y}-{m}-{d}"
            else:
                report_date = datetime.now().strftime("%Y-%m-%d")
    else:
        log_paths, report_date = find_log_sources(args.log_dir, args.date)

    if not log_paths:
        log.error("event=report_failed reason=no_log_files date=%s", report_date or "unknown")
        sys.exit(1)

    log.info("Report date: %s", report_date)

    # Streaming pipeline: read line-by-line, pre-filter zapp records before
    # json.loads, drop out-of-window timestamps without materialising any
    # intermediate list. Peak RAM is dominated by the per-SID accumulator
    # state in build_sessions, not by the raw record count.
    day_start = datetime(*(int(x) for x in report_date.split("-")), tzinfo=tz) \
        .astimezone(ZoneInfo("UTC"))
    day_end = day_start + timedelta(days=1)

    def _date_filtered_records():
        for lp in log_paths:
            n_in = 0
            n_kept = 0
            for rec in iter_log_records(lp):
                n_in += 1
                ts = parse_timestamp(rec.get("TimestampAuthentication", ""))
                if ts and day_start <= ts < day_end:
                    n_kept += 1
                    yield rec
            log.info("  %s: %d zapp records (%d in window)",
                     os.path.basename(lp), n_in, n_kept)

    max_ver = config.max_client_version
    if max_ver > 0:
        log.info("  Version filter: major <= %d", max_ver)
    sessions = build_sessions(_date_filtered_records(), tz, max_client_version=max_ver)
    log.info("  User sessions (after filtering): %d", len(sessions))

    if not sessions:
        log.info(
            "event=report_complete date=%s sessions=0 status=empty mode=%s duration_ms=%d",
            report_date,
            "dry_run" if args.dry_run else "normal",
            int((time.monotonic() - started) * 1000),
        )
        sys.exit(0)

    if args.dry_run:
        log.info(
            "event=report_complete date=%s sessions=%d mode=dry_run "
            "status=ok duration_ms=%d (no files written, no upload attempted)",
            report_date, len(sessions),
            int((time.monotonic() - started) * 1000),
        )
        sys.exit(0)

    # Determine output file — use report_date (from log), not session dates
    filename = config.filename_pattern.replace("{date}", report_date)

    if args.output_file:
        excel_path = args.output_file
    else:
        excel_path = os.path.join(args.output_dir, f"{filename}.xlsx")

    generate_excel(sessions, excel_path)
    log.info("  Excel report: %s", excel_path)

    # Generate CSV report alongside Excel
    csv_path = os.path.splitext(excel_path)[0] + ".csv"
    generate_csv(sessions, csv_path)
    log.info("  CSV report:   %s", csv_path)

    # Generate JSON report alongside Excel
    json_path = os.path.splitext(excel_path)[0] + ".json"
    generate_json(sessions, json_path, args.timezone)
    log.info("  JSON report:  %s", json_path)

    # Lightweight sidecar for the dashboard (avoids loading the full JSON
    # just to render the index or pre-filter search by username).
    summary_path = os.path.splitext(excel_path)[0] + ".summary.json"
    generate_summary_sidecar(
        sessions, summary_path, report_date,
        has_csv=os.path.exists(csv_path),
        has_xlsx=os.path.exists(excel_path),
    )
    log.info("  Summary:      %s", summary_path)

    # Upload to file share if enabled (skip on --no-upload)
    upload_status = "disabled"
    if config.share_enabled and not args.no_upload:
        from share_upload import upload_report
        upload_path = csv_path if config.share_format == "csv" else excel_path
        success, msg = upload_report(upload_path, config)
        upload_status = "ok" if success else "failed"
    elif config.share_enabled and args.no_upload:
        log.info("  Share upload:  skipped (--no-upload)")
        upload_status = "skipped"

    # Clean up old reports
    cleanup_old_reports(args.output_dir, config.retention_days)

    log.info(
        "event=report_complete date=%s sessions=%d excel=%s csv=%s json=%s "
        "upload=%s duration_ms=%d",
        report_date, len(sessions),
        os.path.basename(excel_path), os.path.basename(csv_path),
        os.path.basename(json_path), upload_status,
        int((time.monotonic() - started) * 1000),
    )


if __name__ == "__main__":
    main()
